"""
Serviço de exportação de dados para Excel e PDF.
"""
import json
from pathlib import Path
from typing import List, Dict
from datetime import datetime
import io

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


DATA_PATH = Path(__file__).parent.parent / "data"


class ExportService:
    """Serviço para exportação de dados em diferentes formatos."""

    def __init__(self):
        self.empresas = self._load_json("empresas_exemplo.json")
        self.eventos = self._load_json("eventos.json")
        self.concorrencia = self._load_json("concorrencia.json")

    def _load_json(self, filename: str) -> dict:
        """Carrega arquivo JSON."""
        filepath = DATA_PATH / filename
        if filepath.exists():
            with open(filepath, "r", encoding="utf-8") as f:
                return json.load(f)
        return {}

    def exportar_empresas_excel(self, empresas: List[dict] = None) -> bytes:
        """
        Exporta lista de empresas para Excel.

        Returns:
            Bytes do arquivo Excel
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl não está instalado")

        if empresas is None:
            empresas = self.empresas.get("empresas", [])

        wb = Workbook()
        ws = wb.active
        ws.title = "Empresas Estratégicas"

        # Cabeçalho
        headers = [
            "Nome", "CNPJ", "Setor", "CNAE", "Data Abertura",
            "Município", "Telefone", "Email", "Porte", "Status Parceria"
        ]

        # Estilo do cabeçalho
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Dados
        for row, emp in enumerate(empresas, 2):
            ws.cell(row=row, column=1, value=emp.get("nome") or emp.get("razao_social", ""))
            ws.cell(row=row, column=2, value=emp.get("cnpj", ""))
            ws.cell(row=row, column=3, value=emp.get("setor_hotel") or emp.get("setor", ""))
            ws.cell(row=row, column=4, value=emp.get("cnae_principal") or emp.get("cnae", ""))
            ws.cell(row=row, column=5, value=emp.get("data_abertura", ""))
            ws.cell(row=row, column=6, value=emp.get("municipio", "Ribeirão Pires"))
            ws.cell(row=row, column=7, value=emp.get("telefone", ""))
            ws.cell(row=row, column=8, value=emp.get("email", ""))
            ws.cell(row=row, column=9, value=emp.get("porte", ""))
            ws.cell(row=row, column=10, value=emp.get("status_parceria") or emp.get("status", ""))

        # Ajustar largura das colunas
        column_widths = [30, 18, 20, 12, 12, 15, 15, 25, 8, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Salvar em bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def exportar_relatorio_viabilidade(self) -> bytes:
        """
        Exporta relatório completo de viabilidade para Excel.

        Returns:
            Bytes do arquivo Excel
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl não está instalado")

        wb = Workbook()

        # Aba 1: Resumo
        ws1 = wb.active
        ws1.title = "Resumo Executivo"

        ws1["A1"] = "ESTUDO DE VIABILIDADE - HOTEL RIBEIRÃO PIRES"
        ws1["A1"].font = Font(bold=True, size=16)
        ws1.merge_cells("A1:D1")

        resumo_data = [
            ["", ""],
            ["Data do Relatório", datetime.now().strftime("%d/%m/%Y")],
            ["", ""],
            ["INDICADORES CHAVE", ""],
            ["Visitantes/Ano", "315.000"],
            ["Eventos/Ano", "127"],
            ["Leitos Disponíveis", "200"],
            ["Score de Viabilidade", "82.5"],
            ["", ""],
            ["HOTEL PROPOSTO", ""],
            ["Quartos", "55"],
            ["Diária Target", "R$ 280"],
            ["Categoria", "Upscale"],
            ["", ""],
            ["PROJEÇÃO (Cenário Moderado)", ""],
            ["Ocupação", "60%"],
            ["RevPAR", "R$ 168"],
            ["Receita Anual", "R$ 5.380.000"],
        ]

        for row, (label, value) in enumerate(resumo_data, 3):
            ws1.cell(row=row, column=1, value=label)
            ws1.cell(row=row, column=2, value=value)
            if "INDICADORES" in label or "HOTEL" in label or "PROJEÇÃO" in label:
                ws1.cell(row=row, column=1).font = Font(bold=True)

        # Aba 2: Eventos
        ws2 = wb.create_sheet("Eventos")
        eventos = self.eventos.get("eventos", [])

        headers = ["Evento", "Período", "Público Estimado", "Impacto Hotel"]
        for col, header in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=header).font = Font(bold=True)

        for row, evento in enumerate(eventos, 2):
            ws2.cell(row=row, column=1, value=evento.get("nome", ""))
            ws2.cell(row=row, column=2, value=evento.get("periodo", ""))
            ws2.cell(row=row, column=3, value=evento.get("publico_estimado", 0))
            ws2.cell(row=row, column=4, value=evento.get("impacto_hotel", ""))

        # Aba 3: Concorrência
        ws3 = wb.create_sheet("Concorrência")
        hoteis = self.concorrencia.get("hoteis", [])

        headers = ["Hotel", "Cidade", "Quartos", "Diária Média", "Nota"]
        for col, header in enumerate(headers, 1):
            ws3.cell(row=1, column=col, value=header).font = Font(bold=True)

        for row, hotel in enumerate(hoteis, 2):
            ws3.cell(row=row, column=1, value=hotel.get("nome", ""))
            ws3.cell(row=row, column=2, value=hotel.get("cidade", ""))
            ws3.cell(row=row, column=3, value=hotel.get("quartos", 0))
            ws3.cell(row=row, column=4, value=hotel.get("diaria_media", 0))
            ws3.cell(row=row, column=5, value=hotel.get("nota_avaliacao", ""))

        # Aba 4: Empresas
        ws4 = wb.create_sheet("Empresas")
        empresas = self.empresas.get("empresas", [])

        headers = ["Nome", "Setor", "CNAE", "Status"]
        for col, header in enumerate(headers, 1):
            ws4.cell(row=1, column=col, value=header).font = Font(bold=True)

        for row, emp in enumerate(empresas, 2):
            ws4.cell(row=row, column=1, value=emp.get("nome") or emp.get("razao_social", ""))
            ws4.cell(row=row, column=2, value=emp.get("setor_hotel") or emp.get("setor", ""))
            ws4.cell(row=row, column=3, value=emp.get("cnae_principal") or emp.get("cnae", ""))
            ws4.cell(row=row, column=4, value=emp.get("status_parceria") or emp.get("status", ""))

        # Salvar em bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def gerar_csv_empresas(self, empresas: List[dict] = None) -> str:
        """
        Gera CSV das empresas.

        Returns:
            String CSV
        """
        if empresas is None:
            empresas = self.empresas.get("empresas", [])

        headers = ["Nome", "CNPJ", "Setor", "CNAE", "Telefone", "Email", "Status"]
        lines = [";".join(headers)]

        for emp in empresas:
            row = [
                emp.get("nome") or emp.get("razao_social", ""),
                emp.get("cnpj", ""),
                emp.get("setor_hotel") or emp.get("setor", ""),
                emp.get("cnae_principal") or emp.get("cnae", ""),
                emp.get("telefone", ""),
                emp.get("email", ""),
                emp.get("status_parceria") or emp.get("status", "")
            ]
            lines.append(";".join(str(v) for v in row))

        return "\n".join(lines)


# Instância global
export_service = ExportService()
